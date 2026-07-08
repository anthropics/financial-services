#!/usr/bin/env python3
"""
DDM Model Validation Script
Validates Excel Dividend Discount Model (DDM) workbooks for formula errors
and common DDM mistakes.

Mirrors the conventions of the dcf-model skill's validate_dcf.py so the two
model validators behave and report identically.

Usage:
    python validate_ddm.py <excel_file> [output.json]
"""

import sys
import json
from pathlib import Path


class DDMModelValidator:
    """Validates DDM models for errors and quality issues."""

    def __init__(self, excel_path: str):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        self.excel_path = excel_path
        self.openpyxl = openpyxl

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"File not found: {excel_path}")

        self.workbook_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        self.workbook_values = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors = []
        self.warnings = []
        self.info = []

    def validate_all(self) -> dict:
        """Run all validation checks and return a results dict."""
        from datetime import datetime

        self.check_sheet_structure()
        self.check_formula_errors()
        self.check_ddm_logic()

        return {
            "file": self.excel_path,
            "validation_date": datetime.now().isoformat(),
            "status": "PASS" if len(self.errors) == 0 else "FAIL",
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info,
        }

    # --- structural checks --------------------------------------------------
    def check_sheet_structure(self):
        """Verify recommended sheets exist."""
        recommended = ["DDM", "Sensitivity"]
        sheet_names = self.workbook_values.sheetnames
        for sheet in recommended:
            if sheet not in sheet_names:
                self.warnings.append(f"Recommended sheet missing: {sheet}")
            else:
                self.info.append(f"Found sheet: {sheet}")

    def check_formula_errors(self):
        """Check for Excel formula errors across all sheets."""
        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        total_errors = 0
        total_formulas = 0

        for sheet_name in self.workbook_values.sheetnames:
            ws_values = self.workbook_values[sheet_name]
            ws_formulas = self.workbook_formulas[sheet_name]
            for row in ws_values.iter_rows():
                for cell in row:
                    formula_cell = ws_formulas[cell.coordinate]
                    if (
                        formula_cell.value
                        and isinstance(formula_cell.value, str)
                        and formula_cell.value.startswith("=")
                    ):
                        total_formulas += 1
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                self.errors.append(f"{err} at {location}")
                                total_errors += 1
                                break

        self.info.append(f"Total formulas: {total_formulas}")
        if total_errors == 0:
            self.info.append("✓ No formula errors found")
        else:
            # summary goes in info so error_count reflects the number of erroring cells
            self.info.append(f"Total formula errors: {total_errors}")

    # --- DDM-specific logic -------------------------------------------------
    def check_ddm_logic(self):
        self._check_cost_of_equity_vs_growth()
        self._check_cost_of_equity_range()
        self._check_terminal_growth_reasonable()
        self._check_terminal_value_proportion()
        self._check_payout_ratio()

    def _find_rate_near_label(self, sheet, label_terms, lo=0.0, hi=1.0, exclude_terms=()):
        """Return the first plausible number (lo <= x < hi) adjacent to a cell
        whose text contains ALL of label_terms (and NONE of exclude_terms),
        case-insensitive. Probes cells to the right of the label first, then
        directly below it, so both horizontal and vertical layouts are found."""
        for row in sheet.iter_rows(max_row=200, max_col=25):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    text = cell.value.lower()
                    if all(term in text for term in label_terms) and not any(
                        x in text for x in exclude_terms
                    ):
                        for dr, dc in ((0, 1), (0, 2), (0, 3), (0, 4), (0, 5), (1, 0), (2, 0)):
                            adjacent = sheet.cell(cell.row + dr, cell.column + dc).value
                            if isinstance(adjacent, (int, float)) and lo <= adjacent < hi:
                                return adjacent
        return None

    def _ddm_sheet(self):
        try:
            return self.workbook_values["DDM"]
        except KeyError:
            # fall back to the first sheet so single-sheet models still validate
            return self.workbook_values[self.workbook_values.sheetnames[0]]

    def _cost_of_equity(self, sheet):
        # accept several common labels for the discount rate in a DDM
        for terms in (["cost", "equity"], ["required", "return"], ["discount", "rate"]):
            r = self._find_rate_near_label(sheet, terms)
            if r is not None:
                return r
        return None

    def _terminal_growth(self, sheet):
        for terms in (["terminal", "growth"], ["perpetual", "growth"], ["long-term", "growth"], ["long", "run", "growth"]):
            g = self._find_rate_near_label(sheet, terms)
            if g is not None:
                return g
        return None

    def _check_cost_of_equity_vs_growth(self):
        """CRITICAL: cost of equity must exceed the perpetual growth rate."""
        try:
            sheet = self._ddm_sheet()
            r = self._cost_of_equity(sheet)
            g = self._terminal_growth(sheet)
            if r is not None and g is not None:
                if g >= r:
                    self.errors.append(
                        f"CRITICAL: terminal growth ({g:.2%}) >= cost of equity ({r:.2%}). "
                        "The Gordon terminal value D/(r-g) is infinite or negative and is "
                        "mathematically invalid."
                    )
                else:
                    self.info.append(
                        f"✓ terminal growth ({g:.2%}) < cost of equity ({r:.2%})"
                    )
            else:
                self.warnings.append("Could not locate cost of equity and terminal growth values")
        except Exception as e:  # noqa: BLE001 - report, never crash the run
            self.warnings.append(f"Could not validate cost of equity vs growth: {e}")

    def _check_cost_of_equity_range(self):
        try:
            r = self._cost_of_equity(self._ddm_sheet())
            if r is None:
                self.warnings.append("Could not locate cost of equity value")
            elif r < 0.04 or r > 0.20:
                self.warnings.append(
                    f"Cost of equity ({r:.2%}) is outside the typical 4%-20% range. Verify CAPM inputs."
                )
            else:
                self.info.append(f"✓ Cost of equity ({r:.2%}) in reasonable range")
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Could not validate cost of equity range: {e}")

    def _check_terminal_growth_reasonable(self):
        try:
            g = self._terminal_growth(self._ddm_sheet())
            if g is None:
                self.warnings.append("Could not locate terminal growth value")
            elif g > 0.05:
                self.warnings.append(
                    f"Terminal growth ({g:.2%}) exceeds ~5%. A perpetual rate above long-run "
                    "nominal GDP growth is generally not defensible."
                )
            else:
                self.info.append(f"✓ Terminal growth ({g:.2%}) <= 5%")
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Could not validate terminal growth: {e}")

    def _check_terminal_value_proportion(self):
        """Warn if the PV of the terminal value dominates total value.

        The PV-of-terminal-value and total-value cells must be on the same
        basis (both per-share, or both aggregate). We exclude any candidate
        'total' cell whose label also mentions 'terminal' so the terminal-value
        line cannot be mistaken for the total, and if the resulting ratio is not
        in (0, 1] we treat the two bases as unreconcilable rather than passing
        silently.
        """
        try:
            sheet = self._ddm_sheet()
            tv = self._find_rate_near_label(sheet, ["pv", "terminal"], lo=0, hi=1e15)
            total = self._find_rate_near_label(
                sheet, ["value", "per", "share"], lo=0, hi=1e15, exclude_terms=("terminal",)
            )
            if total is None:
                total = self._find_rate_near_label(
                    sheet, ["intrinsic", "value"], lo=0, hi=1e15, exclude_terms=("terminal",)
                )
            if tv is not None and total is not None and total > 0:
                proportion = tv / total
                if not 0 < proportion <= 1:
                    self.warnings.append(
                        "Could not reconcile PV-of-terminal-value and total-value bases "
                        "(per-share vs aggregate); terminal-value proportion not checked."
                    )
                elif proportion > 0.90:
                    self.warnings.append(
                        f"PV of terminal value is {proportion:.1%} of total value. The model is "
                        "heavily reliant on the perpetuity — consider a longer explicit horizon."
                    )
                else:
                    self.info.append(f"✓ PV of terminal value is {proportion:.1%} of total value")
            else:
                self.warnings.append("Could not locate terminal value and total value")
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Could not validate terminal value proportion: {e}")

    def _check_payout_ratio(self):
        try:
            payout = self._find_rate_near_label(self._ddm_sheet(), ["payout"], lo=0, hi=5)
            if payout is None:
                return  # payout is optional; not every DDM surfaces it
            if payout > 1.0:
                self.warnings.append(
                    f"Payout ratio ({payout:.1%}) exceeds 100% — dividends above earnings are not "
                    "sustainable in perpetuity. Verify the terminal payout assumption."
                )
            else:
                self.info.append(f"✓ Payout ratio ({payout:.1%}) <= 100%")
        except Exception as e:  # noqa: BLE001
            self.warnings.append(f"Could not validate payout ratio: {e}")


def validate_ddm_model(excel_path: str) -> dict:
    """Validate a DDM model Excel file and return a results dict."""
    return DDMModelValidator(excel_path).validate_all()


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_ddm.py <excel_file> [output.json]")
        print("\nValidates a Dividend Discount Model for:")
        print("  - Formula errors (#REF!, #DIV/0!, etc.)")
        print("  - Cost of equity > terminal growth (critical; Gordon requires r > g)")
        print("  - Cost of equity in reasonable range (4-20%)")
        print("  - Terminal growth <= ~5% (long-run nominal GDP)")
        print("  - PV of terminal value not dominating total value")
        print("  - Terminal payout ratio <= 100%")
        print("\nReturns JSON with errors, warnings, and info.")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        results = validate_ddm_model(excel_file)
        print(json.dumps(results, indent=2))
        if output_file:
            with open(output_file, "w") as f:
                json.dump(results, f, indent=2)
        sys.exit(0 if results["status"] == "PASS" else 1)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"file": excel_file, "status": "ERROR", "error": str(e)}, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
