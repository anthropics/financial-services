#!/usr/bin/env python3
"""
News Sentiment Brief Validation Script
Validates the Excel "Sentiment" table produced by the news-sentiment skill for
schema integrity and the skill's non-negotiable rules: every item carries a
verbatim source quote, sentiment uses the fixed label set, any numeric sentiment
score is in range and consistent with its label, and low-confidence / mixed
items are flagged for human review.

Mirrors the conventions of the dcf-model validator (class-based, errors /
warnings / info, JSON output, matching exit codes).

Usage:
    python validate_sentiment.py <excel_file> [output.json]
"""

import sys
import json
from pathlib import Path

# One shared threshold partitions the score axis into disjoint, exhaustive
# regions for the non-"mixed" labels, matching references/labeling-guide.md:
#   neutral iff |score| <= NEUTRAL_BAND ; positive iff score > NEUTRAL_BAND ;
#   negative iff score < -NEUTRAL_BAND. "mixed" is a net figure read from the
#   label, not the band, and is excluded from these checks.
NEUTRAL_BAND = 0.15
MIXED_MAX_ABS = 0.5          # a |net score| above this suggests pos/neg, not mixed
CONF_LOW_CUTOFF = 0.34       # numeric confidence at/below this counts as "low"

SENTIMENT_LABELS = {"positive", "neutral", "negative", "mixed"}
CONFIDENCE_LABELS = {"low", "medium", "med", "high"}
EVENT_TAXONOMY = {
    "", "earnings", "guidance", "m&a", "ma", "merger", "acquisition", "litigation",
    "legal", "management", "product", "regulatory", "regulation", "macro",
    "capital-return", "capitalreturn", "buyback", "dividend",
    "rating-change", "ratingchange", "rating", "other",
}

# header synonym -> canonical field
FIELD_SYNONYMS = {
    "id": {"id", "item_id", "item", "no", "num", "#"},
    "source": {"source", "publisher", "outlet", "document", "doc"},
    "date": {"date", "datetime", "published", "publisheddate"},
    "sentiment": {"sentiment", "label", "stance"},
    "source_quote": {"source_quote", "sourcequote", "quote", "evidence", "excerpt"},
    "sentiment_score": {"sentiment_score", "sentimentscore", "score", "polarity"},
    "confidence": {"confidence", "conf"},
    "event_type": {"event_type", "eventtype", "event", "category", "tag", "eventtag", "eventtags"},
    "entities": {"entities", "entity", "tickers"},
    "needs_review": {"needs_review", "needsreview", "review", "flag", "flagged"},
}
REQUIRED_FIELDS = ("id", "source", "sentiment", "source_quote")
TRUTHY = {"true", "yes", "y", "1", "x", "flag", "flagged"}


def _norm(s) -> str:
    return "".join(str(s).strip().lower().split()).replace("-", "_") if s is not None else ""


class SentimentBriefValidator:
    def __init__(self, excel_path: str):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"File not found: {excel_path}")

        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors = []
        self.warnings = []
        self.info = []

    def validate_all(self) -> dict:
        from datetime import datetime

        sheet = self._sentiment_sheet()
        if sheet is not None:
            cols = self._map_columns(sheet)
            if cols is not None:
                self._validate_rows(sheet, cols)

        return {
            "file": self.excel_path,
            "validation_date": datetime.now().isoformat(),
            "status": "PASS" if len(self.errors) == 0 else "FAIL",
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info,
        }

    def _sentiment_sheet(self):
        for name in self.wb.sheetnames:
            if "sentiment" in name.lower():
                self.info.append(f"Found sheet: {name}")
                return self.wb[name]
        if self.wb.sheetnames:
            self.warnings.append(
                "No sheet named 'Sentiment' found; validating the first sheet instead."
            )
            return self.wb[self.wb.sheetnames[0]]
        self.errors.append("Workbook has no sheets.")
        return None

    def _map_columns(self, sheet):
        """Map the header row (row 1) to canonical field names -> column index."""
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header or all(c is None for c in header):
            self.errors.append("Header row is empty; expected column titles in row 1.")
            return None

        cols = {}
        for idx, cell in enumerate(header):
            n = _norm(cell)
            for field, syns in FIELD_SYNONYMS.items():
                if n in syns and field not in cols:
                    cols[field] = idx
        missing = [f for f in REQUIRED_FIELDS if f not in cols]
        if missing:
            self.errors.append(
                f"Missing required column(s): {', '.join(missing)} "
                f"(found: {', '.join(sorted(cols)) or 'none'})."
            )
            return None
        self.info.append(f"Columns mapped: {', '.join(sorted(cols))}")
        return cols

    def _confidence_is_low(self, conf) -> bool:
        """True iff confidence clearly denotes low confidence (enum 'low' or numeric <= cutoff)."""
        if conf is None or str(conf).strip() == "":
            return False
        cn = _norm(conf)
        if cn == "low":
            return True
        try:
            return float(conf) <= CONF_LOW_CUTOFF
        except (TypeError, ValueError):
            return False

    def _validate_rows(self, sheet, cols):
        n_rows = 0
        dist = {k: 0 for k in SENTIMENT_LABELS}
        review_flags = 0
        has_review_col = "needs_review" in cols
        unflagged_needing_review = 0  # only meaningful when has_review_col is False

        for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            n_rows += 1

            def get(field):
                i = cols.get(field)
                return row[i] if (i is not None and i < len(row)) else None

            sent = _norm(get("sentiment"))
            if sent not in SENTIMENT_LABELS:
                self.errors.append(
                    f"Row {r}: sentiment '{get('sentiment')}' is not one of {sorted(SENTIMENT_LABELS)}."
                )
            else:
                dist[sent] += 1

            # source quote (required, non-empty — the auditability rule)
            quote = get("source_quote")
            if quote is None or str(quote).strip() == "":
                self.errors.append(f"Row {r}: source_quote is empty — every label needs a verbatim quote.")

            # numeric score: range is enforced (error); label consistency is advisory (warning)
            if "sentiment_score" in cols:
                raw = get("sentiment_score")
                if raw is not None and str(raw).strip() != "":
                    try:
                        score = float(raw)
                        if not -1.0 <= score <= 1.0:
                            self.errors.append(f"Row {r}: sentiment_score {score} outside [-1, 1].")
                        elif sent == "mixed":
                            if abs(score) > MIXED_MAX_ABS:
                                self.warnings.append(
                                    f"Row {r}: 'mixed' label but |score| {abs(score):.2f} > {MIXED_MAX_ABS} — "
                                    "an item this one-sided is probably positive/negative, not mixed."
                                )
                        elif sent == "positive" and score <= NEUTRAL_BAND:
                            self.warnings.append(f"Row {r}: 'positive' label but score {score} <= {NEUTRAL_BAND}.")
                        elif sent == "negative" and score >= -NEUTRAL_BAND:
                            self.warnings.append(f"Row {r}: 'negative' label but score {score} >= -{NEUTRAL_BAND}.")
                        elif sent == "neutral" and abs(score) > NEUTRAL_BAND:
                            self.warnings.append(f"Row {r}: 'neutral' label but |score| {abs(score):.2f} > {NEUTRAL_BAND}.")
                    except (TypeError, ValueError):
                        self.warnings.append(f"Row {r}: sentiment_score '{raw}' is not numeric.")

            # confidence (optional): enum or 0 < x <= 1
            conf = get("confidence") if "confidence" in cols else None
            if "confidence" in cols and conf is not None and str(conf).strip() != "":
                cn = _norm(conf)
                ok = cn in CONFIDENCE_LABELS
                if not ok:
                    try:
                        f = float(conf)
                        ok = 0.0 < f <= 1.0
                    except (TypeError, ValueError):
                        ok = False
                if not ok:
                    self.warnings.append(
                        f"Row {r}: confidence '{conf}' is not low/medium/high or a value in (0, 1]."
                    )

            # needs_review coupling: low-confidence OR mixed items MUST be flagged
            requires_review = sent == "mixed" or self._confidence_is_low(conf)
            if has_review_col:
                flagged = _norm(get("needs_review")) in TRUTHY
                if flagged:
                    review_flags += 1
                if requires_review and not flagged:
                    reason = "mixed" if sent == "mixed" else "low-confidence"
                    self.warnings.append(
                        f"Row {r}: {reason} item is not flagged needs_review=yes — it must be queued for human review."
                    )
            elif requires_review:
                unflagged_needing_review += 1

            # event type (optional): taxonomy (warning only — extensible)
            if "event_type" in cols:
                ev = get("event_type")
                if ev is not None and str(ev).strip() != "":
                    for token in str(ev).replace(";", ",").split(","):
                        if _norm(token).replace("_", "-") not in EVENT_TAXONOMY and _norm(token) not in EVENT_TAXONOMY:
                            self.warnings.append(
                                f"Row {r}: event_type '{token.strip()}' not in the taxonomy (use 'other' if unsure)."
                            )

        if n_rows == 0:
            self.warnings.append("No data rows found under the header.")
            return

        self.info.append(f"Validated {n_rows} item(s).")
        self.info.append("Sentiment distribution: " + ", ".join(f"{k}={v}" for k, v in dist.items() if v))
        if has_review_col:
            self.info.append(f"Items flagged for review: {review_flags}")
        elif unflagged_needing_review:
            self.warnings.append(
                f"{unflagged_needing_review} low-confidence/mixed item(s) require review but there is no "
                "needs_review column to flag them — add one."
            )


def validate_sentiment_brief(excel_path: str) -> dict:
    return SentimentBriefValidator(excel_path).validate_all()


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_sentiment.py <excel_file> [output.json]")
        print("\nValidates a News Sentiment Brief workbook for:")
        print("  - A 'Sentiment' sheet with required columns (id, source, sentiment, source_quote)")
        print("  - sentiment restricted to positive/neutral/negative/mixed")
        print("  - a non-empty verbatim source_quote on every item")
        print("  - sentiment_score in [-1, 1] (error) and consistent with the label (warning)")
        print("  - low-confidence / mixed items flagged needs_review=yes (warning)")
        print("  - confidence and event_type well-formed (warnings)")
        print("\nReturns JSON with errors, warnings, and info.")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    try:
        results = validate_sentiment_brief(excel_file)
        print(json.dumps(results, indent=2))
        if output_file:
            with open(output_file, "w") as f:
                json.dump(results, f, indent=2)
        sys.exit(0 if results["status"] == "PASS" else 1)
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"file": excel_file, "status": "ERROR", "error": str(e)}, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
